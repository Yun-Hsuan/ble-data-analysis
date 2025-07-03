import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, numbers

def generate_tenant_comparison(overview_path: str, output_path: str):
    """
    從 Overview Report 生成櫃位比較報表。
    
    :param overview_path: Overview Report 的路徑
    :param output_path: 輸出報表的路徑
    """
    # 定義要比較的指標
    indicators = [
        "路過數 A", "靠櫃數 B", "靠櫃率 B/A", 
        "停留數 C", "停留率 C/B", 
        "交易數 D", "提袋率 D/C", "提袋率 D/B", 
        "平均停留 (秒)"
    ]
    
    # 讀取 Overview Report
    print(f"讀取 Overview Report: {overview_path}")
    excel_data = pd.ExcelFile(overview_path)
    
    # 初始化結果 DataFrame
    results = pd.DataFrame(index=indicators)
    
    # 處理每個櫃位（每個工作表）
    for sheet_name in excel_data.sheet_names:
        print(f"處理櫃位: {sheet_name}")
        # 讀取該櫃位的資料
        df = excel_data.parse(sheet_name)
        
        # 計算每個指標的平均值
        averages = {}
        for indicator in indicators:
            averages[indicator] = df[indicator].mean()
        
        # 將結果加入到結果 DataFrame
        results[sheet_name] = pd.Series(averages)
    
    # 寫入 Excel
    print(f"生成比較報表: {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        results.to_excel(writer, sheet_name="櫃位比較")
        
        # 設定百分比格式
        worksheet = writer.sheets["櫃位比較"]
        for col in range(2, worksheet.max_column + 1):  # 跳過索引列
            for row in range(2, worksheet.max_row + 1):  # 跳過標題列
                cell = worksheet.cell(row=row, column=col)
                if any(rate in worksheet.cell(row=row, column=1).value for rate in ["率"]):
                    cell.number_format = "0.00%"
    
    print("完成！")

if __name__ == "__main__":
    generate_tenant_comparison(
        overview_path="./daily_reports/Overview_Report.xlsx",
        output_path="./daily_reports/tenant_comparison.xlsx"
    ) 