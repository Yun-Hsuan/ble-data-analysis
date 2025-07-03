import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, numbers
import pandas as pd
from datetime import datetime, timedelta
from src.data_processing.cleaners.ble_cleaner import BLECleaner
from src.data_processing.cleaners.transaction_cleaner import TransactionCleaner
from src.business.tenant_indicators.visit_rate import VisitRateIndicator
from src.business.tenant_indicators.pass_by import PassByIndicator
from src.business.tenant_indicators.dwell_rate import DwellRateIndicator
from src.business.tenant_indicators.bagging_rate import BaggingRateIndicator
from src.report.report_manager import ReportManager

def generate_overview_report(input_dir: str, output_path: str):
    """
    Generate an overview report comparing all tenants' indicators across multiple days and floors.
    :param input_dir: Directory containing daily report Excel files.
    :param output_path: Path to save the overview report.
    """
    input_dir = Path(input_dir)
    report_files = sorted(input_dir.glob("tenant_report_*.xlsx"))

    # 初始化儲存所有櫃位的資料
    all_tenants_data = {}

    print(f"Found report files: {report_files}")

    # 處理每個每日報告檔案
    for report_file in report_files:
        # 從檔案名稱中提取日期和樓層資訊
        file_name = report_file.stem
        date = file_name.split("_")[-2]  # 提取日期
        floor = file_name.split("_")[-1]  # 提取樓層
        
        print(f"Processing file: {report_file}")
        excel_data = pd.ExcelFile(report_file)

        # 處理每個工作表（櫃位）
        for sheet_name in excel_data.sheet_names:
            df = excel_data.parse(sheet_name)
            total_row = df[df["日期"] == "Total"]  # 過濾 'Total' 列

            if not total_row.empty:
                metrics = {
                    "日期": date,
                    "樓層": floor,
                    "櫃位": sheet_name,
                    "路過數 A": total_row["行經數 A"].values[0],
                    "靠櫃數 B": total_row["入櫃數 B"].values[0],
                    "靠櫃率 B/A": total_row["入櫃率 B/A"].values[0],
                    "停留數 C": total_row["停留數 C"].values[0],
                    "停留率 C/B": total_row["停留率 C/B"].values[0],
                    "交易數 D": total_row["交易數 D"].values[0],
                    "提袋率 D/C": total_row["提袋率 D/C"].values[0],
                    "提袋率 D/B": total_row["提袋率 D/B"].values[0],
                    "平均停留 (秒)": total_row["平均停留時長 (s)"].values[0],
                }
                
                # 將資料加入對應櫃位的列表中
                if sheet_name not in all_tenants_data:
                    all_tenants_data[sheet_name] = []
                all_tenants_data[sheet_name].append(metrics)

    # 建立一個百分比樣式
    percent_style = NamedStyle(name="percent_style", number_format=numbers.FORMAT_PERCENTAGE_00)

    # 寫入總覽 Excel 檔案
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 寫入每個櫃位的資料
        for tenant_name, tenant_data in all_tenants_data.items():
            tenant_df = pd.DataFrame(tenant_data)
            
            # 檢查重複資料
            duplicates = tenant_df.duplicated(subset=['日期', '樓層', '櫃位'], keep=False)
            if duplicates.any():
                print(f"\n發現重複資料 - 櫃位: {tenant_name}")
                print(tenant_df[duplicates])
                
                # 對重複資料進行聚合（取平均）
                tenant_df = tenant_df.groupby(['日期', '樓層', '櫃位']).agg({
                    '路過數 A': 'mean',
                    '靠櫃數 B': 'mean',
                    '靠櫃率 B/A': 'mean',
                    '停留數 C': 'mean',
                    '停留率 C/B': 'mean',
                    '交易數 D': 'mean',
                    '提袋率 D/C': 'mean',
                    '提袋率 D/B': 'mean',
                    '平均停留 (秒)': 'mean'
                }).reset_index()
                
                print(f"已處理重複資料，使用平均值")
            
            # 按照日期和樓層排序
            tenant_df = tenant_df.sort_values(["日期", "樓層"])
            
            # 寫入工作表
            sheet_name = tenant_name.replace("/", "_")  # 處理可能的特殊字元
            tenant_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 應用百分比格式
            worksheet = writer.sheets[sheet_name]
            
            # 要格式化的百分比欄位
            percentage_columns = ["靠櫃率 B/A", "停留率 C/B", "提袋率 D/C", "提袋率 D/B"]
            column_indices = [tenant_df.columns.get_loc(col) + 1 for col in percentage_columns if col in tenant_df.columns]

            for col_idx in column_indices:
                column_letter = get_column_letter(col_idx)
                for row in range(2, worksheet.max_row + 1):  # 跳過標題列
                    worksheet[f"{column_letter}{row}"].style = percent_style

    print(f"\nOverview report successfully generated: {output_path}")

def date_range(start_date: str, end_date: str):
    """
    Generate a range of dates between start_date and end_date (inclusive).
    :param start_date: Start date in the format 'YYYY-MM-DD'.
    :param end_date: End date in the format 'YYYY-MM-DD'.
    :return: A list of date strings in the format 'YYYY-MM-DD'.
    """
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    return [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range((end - start).days + 1)]

if __name__ == "__main__":

    #Example usage
    # generate_overview_report(
    #     input_dir="./daily_reports",
    #     output_path="./daily_reports/NikeAdidas_Overview_Report.xlsx"
    # )

    #exit()

    date_list = date_range("2025-05-12", "2025-05-18")
    for date in date_list:
        print(f"Processing date: {date}")

        ble_cleaner = BLECleaner(directory=f"data/processed/ble")
        transaction_cleaner = TransactionCleaner(directory=f"data/intermediate/transactions")
        print("Loading ble data...")
        ble_cleaner.load_data(pattern=fr"{date}.*\.csv")
        transaction_cleaner.load_data(pattern=fr"{date}.*\.csv")

        pass_by_indicator = PassByIndicator()
        pass_by_indicator.set_cleaner("ble_cleaner", ble_cleaner)
        pass_by_indicator.set_cleaner("transaction_cleaner", transaction_cleaner)
        pass_by_indicator.set_tenant_mapping("./data/processed/tenant_info/tenant_terminalId_mappingtable_nancy34.xlsx")
        pass_by_indicator.set_rssi_thresholds_from_file("./data/processed/tenant_info/tenant_rssi_thresholds_nancy34.xlsx")

        visit_indicator = VisitRateIndicator()
        visit_indicator.set_cleaner("ble_cleaner", ble_cleaner)
        visit_indicator.set_cleaner("transaction_cleaner", transaction_cleaner)
        visit_indicator.set_tenant_mapping("./data/processed/tenant_info/tenant_terminalId_mappingtable_nancy34.xlsx")
        visit_indicator.set_rssi_thresholds_from_file("./data/processed/tenant_info/tenant_rssi_thresholds_nancy34.xlsx")

        dwell_indicator = DwellRateIndicator([60])
        dwell_indicator.set_cleaner("ble_cleaner", ble_cleaner)
        dwell_indicator.set_cleaner("transaction_cleaner", transaction_cleaner)
        dwell_indicator.set_tenant_mapping("./data/processed/tenant_info/tenant_terminalId_mappingtable_nancy34.xlsx")
        dwell_indicator.set_rssi_thresholds_from_file("./data/processed/tenant_info/tenant_rssi_thresholds_nancy34.xlsx")

        bagging_indicator = BaggingRateIndicator()
        bagging_indicator.set_cleaner("ble_cleaner", ble_cleaner)
        bagging_indicator.set_cleaner("transaction_cleaner", transaction_cleaner)
        bagging_indicator.set_tenant_mapping("./data/processed/tenant_info/tenant_terminalId_mappingtable_nancy34.xlsx")
        bagging_indicator.set_rssi_thresholds_from_file("./data/processed/tenant_info/tenant_rssi_thresholds_nancy34.xlsx")

        report_manager = ReportManager(output_dir="./daily_reports2", tenant_mapping_path="./data/processed/tenant_info/tenant_terminalId_mappingtable_nancy34.xlsx")

        report_manager.generate_reports_for_date_range(
            start_date=date,
            end_date=date,
            pass_by_indicator=pass_by_indicator,
            visit_indicator=visit_indicator,
            dwell_indicator=dwell_indicator,
            bagging_indicator=bagging_indicator
        )