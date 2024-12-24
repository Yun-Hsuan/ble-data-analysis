import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, numbers
import json
import pandas as pd
from datetime import datetime, timedelta
from src.data_processing.cleaners.ble_cleaner import BLECleaner
from src.data_processing.cleaners.transaction_cleaner import TransactionCleaner
from src.business.tenant_indicators.visit_rate import VisitRateIndicator
from src.business.tenant_indicators.pass_by import PassByIndicator
from src.business.tenant_indicators.dwell_rate import DwellRateIndicator
from src.business.tenant_indicators.bagging_rate import BaggingRateIndicator
from src.report.report_manager import ReportManager

def generate_overview_report(input_dir: str, output_path: str):
    """
    Generate an overview report comparing Nike and Adidas indicators across multiple days.
    :param input_dir: Directory containing daily report Excel files.
    :param output_path: Path to save the overview report.
    """
    input_dir = Path(input_dir)
    report_files = sorted(input_dir.glob("tenant_report_*.xlsx"))

    # Initialize storage for aggregated data
    nike_data = []
    adidas_data = []

    print(report_files)

    # Process each daily report file
    for report_file in report_files:
        date = report_file.stem.split("_")[-1]  # Extract date from file name
        excel_data = pd.ExcelFile(report_file)

        for sheet_name in ["NIKE", "ADIDAS"]:
            df = excel_data.parse(sheet_name)
            total_row = df[df["日期"] == "Total"]  # Filter the 'Total' row

            if not total_row.empty:
                metrics = {
                    "日期": date,
                    "路過數 A": total_row["行經數 A"].values[0],
                    "靠櫃數 B": total_row["入櫃數 B"].values[0],
                    "靠櫃率 B/A": total_row["入櫃率 B/A"].values[0],
                    "停留數 C": total_row["停留數 C"].values[0],
                    "停留率 C/B": total_row["停留率 C/B"].values[0],
                    "交易數 D": total_row["交易數 D"].values[0],
                    "提袋率 D/C": total_row["提袋率 D/C"].values[0],
                    "提袋率 D/B": total_row["提袋率 D/B"].values[0],
                    # "平均停留 (秒)": total_row["平均停留時長 (s)"].values[0],
                }
                if sheet_name == "NIKE":
                    nike_data.append(metrics)
                elif sheet_name == "ADIDAS":
                    adidas_data.append(metrics)

    # Create DataFrames for Nike and Adidas
    nike_df = pd.DataFrame(nike_data)
    adidas_df = pd.DataFrame(adidas_data)

    # Create a percentage style for openpyxl
    percent_style = NamedStyle(name="percent_style", number_format=numbers.FORMAT_PERCENTAGE_00)

    # Write to an overview Excel file
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        nike_df.to_excel(writer, sheet_name="Nike_Overview", index=False)
        adidas_df.to_excel(writer, sheet_name="Adidas_Overview", index=False)

        # Apply percentage formatting
        workbook = writer.book
        for sheet_name, df in zip(["Nike_Overview", "Adidas_Overview"], [nike_df, adidas_df]):
            worksheet = workbook[sheet_name]

            # Columns to format as percentage
            percentage_columns = ["靠櫃率 B/A", "停留率 C/B", "提袋率 D/C", "提袋率 D/B"]
            column_indices = [df.columns.get_loc(col) + 1 for col in percentage_columns if col in df.columns]

            for col_idx in column_indices:
                column_letter = get_column_letter(col_idx)
                for row in range(2, worksheet.max_row + 1):  # Skip header row
                    worksheet[f"{column_letter}{row}"].style = percent_style

    print(f"Overview report successfully generated: {output_path}")

def date_range(start_date: str, end_date: str):
    """
    Generate a range of dates between start_date and end_date (inclusive).
    :param start_date: Start date in the format 'YYYY-MM-DD'.
    :param end_date: End date in the format 'YYYY-MM-DD'.
    :return: A list of date strings in the format 'YYYY-MM-DD'.
    """
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    return [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range((end - start).days + 1)]

if __name__ == "__main__":

    #Example usage
    generate_overview_report(
        input_dir="./daily_reports",
        output_path="./daily_reports/NikeAdidas_Overview_Report.xlsx"
    )

    exit()

    date_list = date_range("2024-12-19", "2024-12-22")
    for date in date_list:
        print(f"Processing date: {date}")

        ble_cleaner = BLECleaner(directory=f"data/processed/ble")
        transaction_cleaner = TransactionCleaner(directory=f"data/intermediate/transactions")
        print("Loading ble data...")
        ble_cleaner.load_data(pattern=fr"{date}.*\.csv")
        transaction_cleaner.load_data(pattern=fr"{date}.*\.csv")

        pass_by_indicator = PassByIndicator()
        pass_by_indicator.set_cleaner("ble_cleaner", ble_cleaner)
        pass_by_indicator.set_cleaner("transaction_cleaner", transaction_cleaner)
        pass_by_indicator.set_tenant_mapping("./data/processed/tenant_info/tenant_terminalId_mappingtable.xlsx")
        pass_by_indicator.set_rssi_thresholds_from_file("./data/processed/tenant_info/tenant_rssi_thresholds.xlsx")

        visit_indicator = VisitRateIndicator()
        visit_indicator.set_cleaner("ble_cleaner", ble_cleaner)
        visit_indicator.set_cleaner("transaction_cleaner", transaction_cleaner)
        visit_indicator.set_tenant_mapping("./data/processed/tenant_info/tenant_terminalId_mappingtable.xlsx")
        visit_indicator.set_rssi_thresholds_from_file("./data/processed/tenant_info/tenant_rssi_thresholds.xlsx")

        dwell_indicator = DwellRateIndicator([60])
        dwell_indicator.set_cleaner("ble_cleaner", ble_cleaner)
        dwell_indicator.set_cleaner("transaction_cleaner", transaction_cleaner)
        dwell_indicator.set_tenant_mapping("./data/processed/tenant_info/tenant_terminalId_mappingtable.xlsx")
        dwell_indicator.set_rssi_thresholds_from_file("./data/processed/tenant_info/tenant_rssi_thresholds.xlsx")

        bagging_indicator = BaggingRateIndicator()
        bagging_indicator.set_cleaner("ble_cleaner", ble_cleaner)
        bagging_indicator.set_cleaner("transaction_cleaner", transaction_cleaner)
        bagging_indicator.set_tenant_mapping("./data/processed/tenant_info/tenant_terminalId_mappingtable.xlsx")
        bagging_indicator.set_rssi_thresholds_from_file("./data/processed/tenant_info/tenant_rssi_thresholds.xlsx")

        report_manager = ReportManager(output_dir="./daily_reports", tenant_mapping_path="./data/processed/tenant_info/tenant_terminalId_mappingtable.xlsx")

        report_manager.generate_reports_for_date_range(
            start_date=date,
            end_date=date,
            pass_by_indicator=pass_by_indicator,
            visit_indicator=visit_indicator,
            dwell_indicator=dwell_indicator,
            bagging_indicator=bagging_indicator
        )