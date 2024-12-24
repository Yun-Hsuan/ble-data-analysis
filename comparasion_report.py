import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, numbers
import json
from typing import Optional, Dict, Tuple, List
import os
import pandas as pd
from datetime import datetime, timedelta
from src.data_processing.cleaners.ble_cleaner import BLECleaner
from src.data_processing.cleaners.transaction_cleaner import TransactionCleaner
from src.business.tenant_indicators.visit_rate import VisitRateIndicator
from src.business.tenant_indicators.pass_by import PassByIndicator
from src.business.tenant_indicators.dwell_rate import DwellRateIndicator
from src.business.tenant_indicators.bagging_rate import BaggingRateIndicator
from src.report.report_manager import ReportManager

def generate_comparison_overview_report(comparison_reports_dir: str, overview_report_path: str):
    """
    Generate an overview report comparing all tenants across different indicators.

    :param comparison_reports_dir: Directory containing comparison report Excel files.
    :param overview_report_path: Path to save the generated overview report.
    """
    # Find all comparison report files
    comparison_files = [
        f for f in os.listdir(comparison_reports_dir) if f.startswith("tenant_comparison_") and f.endswith(".xlsx")
    ]

    if not comparison_files:
        print("No comparison report files found.")
        return

    # Initialize a dictionary to store aggregated data for each indicator
    aggregated_data: Dict[str, pd.DataFrame] = {}

    for file_name in comparison_files:
        file_path = os.path.join(comparison_reports_dir, file_name)
        workbook = load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            # Load the sheet into a DataFrame
            sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)

            # Check if "Total" exists in the index
            if "Total" not in sheet_data.index:
                print(f"Skipping {sheet_name} in {file_name}: No Total row found.")
                continue

            # Extract the "Total" row and add it to the aggregated data
            total_row = sheet_data.loc[["Total"]]
            total_row.insert(0, "Date", file_name.split("_")[2].replace(".xlsx", ""))

            if sheet_name not in aggregated_data:
                aggregated_data[sheet_name] = total_row
            else:
                aggregated_data[sheet_name] = pd.concat([aggregated_data[sheet_name], total_row], ignore_index=True)

    # Write aggregated data to the overview report
    with pd.ExcelWriter(overview_report_path, engine="openpyxl") as writer:
        for indicator, df in aggregated_data.items():
            sheetname = indicator.replace("/", "_")  # Replace '/' with '_'
            df.to_excel(writer, sheet_name=indicator, index=False)

            if "率" in sheetname:
                workbook = writer.book
                worksheet = writer.sheets[sheetname]
                for row_idx in range(2, len(df) + 2):  # Exclude header row
                    for col_idx in range(2, len(df.columns) + 2):  # Skip index column
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = "0.00%"

    print(f"Overview report generated: {overview_report_path}")


if __name__ == "__main__":

    generate_comparison_overview_report(
        comparison_reports_dir="./comparison_reports",
        overview_report_path="./comparison_reports/comparison_reports_overview.xlsx"
    )
    exit()

    start_date = "2024-12-09"
    end_date = "2024-12-22"

    # 將日期字串轉換為 datetime 對象
    start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
    end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")

    # 日期迴圈，生成多天的比較報表
    current_date = start_date_obj
    while current_date <= end_date_obj:
        # 生成當前日期的字串
        date_str = current_date.strftime("%Y-%m-%d")
        comparison_report_path = f"./comparison_reports/tenant_comparison_{date_str}.xlsx"

        report_manager = ReportManager(
            daily_reports_dir="./daily_reports",  # 日報表的目錄
            comparison_report_path=comparison_report_path  # 每日報表的路徑稍後設置
        )

        # 載入日報表的資料
        tenant_data = report_manager.load_tenant_data(start_date=date_str, end_date=date_str)

        # 生成比較報表
        report_manager.generate_comparison_report(tenant_data=tenant_data)

        # 移動到下一天
        current_date += timedelta(days=1)

    print("所有比較報表已生成完成。")